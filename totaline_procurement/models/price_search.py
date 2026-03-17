# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import requests
import json
import re
import math
import base64
import logging
from difflib import SequenceMatcher
from markupsafe import escape as html_escape
import io

_logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel parsing
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning('openpyxl not installed. Excel import will not work.')

# ============================================
# CONSTANTS (ported from n8n workflow)
# ============================================

TRUSTED_STORES = [
    "trendyol", "hepsiburada", "amazon.com.tr", "amazon", "migros",
    "migros sanalmarket", "a101", "bim", "sok", "carrefour",
    "carrefoursa", "gratis", "watsons", "avansas", "ofix",
    "bizim toptan", "metro", "makro", "koctas", "bauhaus",
    "n11", "gittigidiyor", "ciceksepeti", "pttavm", "morhipo",
    "teknosa", "mediamarkt", "vatan", "getir", "istegelsin",
    "rossmann", "eve", "flo", "boyner", "lcw", "defacto",
    "walmart", "costco", "sanalmarket", "marketpaketi", "happy center",
    # B2B / Wholesale platforms
    "toptancidan", "toptan perakende", "toptanburada", "uygunmarket",
    "cimri", "akakce", "epey",
]

ACCESSORY_KEYWORDS = [
    # Cups and drinkware
    'bardak', 'bardagi', 'cup', 'mug', 'fincan',
    # Gifts and extras
    'hediye', 'hediyeli', 'gift', 'bonus', 'promosyon',
    # Accessories
    'kasik', 'spoon', 'tabak', 'plate',
    # Combo indicators
    '+ ', ' + ', 'ile birlikte', 'set icerigi', 'kullan at',
    # Free items
    'ucretsiz', 'bedava', 'free', 'gratis',
]

# Generic product names that need description-based search
GENERIC_PRODUCT_NAMES = [
    'rulo pecete', 'rulo pecete', 'kagit havlu', 'kagit havlu',
    'pecete', 'pecete', 'havlu',
]

# Map generic product names to better search terms
# "Rulo Peçete" is colloquial — "kağıt havlu" is what Google Shopping indexes
GENERIC_NAME_MAP = {
    'rulo pecete': 'kağıt havlu',
    'kagit havlu': 'kağıt havlu',
    'havlu': 'kağıt havlu',
}

# Product keyword categories for relevance checking
PRODUCT_KEYWORDS = {
    'turk kahvesi': ['turk kahvesi', 'turk kahvesi', 'turkish coffee'],
    'filtre kahve': ['filtre kahve', 'filter coffee'],
    'cay': ['cay', 'cay', 'tea'],
    'deterjan': ['deterjan', 'tablet', 'detergent'],
    'kagit havlu': ['kagit havlu', 'kagit havlu', 'paper towel', 'rulo'],
    'islak mendil': ['islak', 'mendil', 'havlu', 'wet wipe'],
}

# Wrong product categories to penalize
WRONG_CATEGORIES = ['bebek bezi', 'kedi mamasi', 'kopek mamasi', 'oyuncak']

# Turkish character mapping for normalization
TURKISH_CHAR_MAP = {
    '\u0131': 'i',  # ı → i
    '\u011f': 'g',  # ğ → g
    '\u00fc': 'u',  # ü → u
    '\u015f': 's',  # ş → s
    '\u00f6': 'o',  # ö → o
    '\u00e7': 'c',  # ç → c
    '\u0130': 'i',  # İ → i
    '\u011e': 'g',  # Ğ → g
    '\u00dc': 'u',  # Ü → u
    '\u015e': 's',  # Ş → s
    '\u00d6': 'o',  # Ö → o
    '\u00c7': 'c',  # Ç → c
}


# ============================================
# HELPER FUNCTIONS (ported from n8n JS)
# ============================================

def normalize_turkish(text):
    """Normalize Turkish characters to ASCII equivalents."""
    if not text:
        return ''
    result = str(text).lower()
    for turkish_char, ascii_char in TURKISH_CHAR_MAP.items():
        result = result.replace(turkish_char, ascii_char)
    return result


def fuzzy_brand_match(brand_word, title_words, threshold=0.75):
    """Check if brand_word fuzzy-matches any word in title_words."""
    for tw in title_words:
        if len(tw) < 2:
            continue
        # Exact substring (fast path)
        if brand_word in tw or tw in brand_word:
            return True
        # Fuzzy match for words of similar length
        if abs(len(brand_word) - len(tw)) <= 2:
            ratio = SequenceMatcher(None, brand_word, tw).ratio()
            if ratio >= threshold:
                return True
    return False


def clean_description(description):
    """Clean description text for use in search queries.
    Removes parenthetical math, calculation expressions, and noise.
    """
    if not description:
        return ''
    text = description.strip()
    # Remove parenthetical expressions containing math: (15 aket 15*8=120 adet)
    text = re.sub(r'\([^)]*[*=+/][^)]*\)', '', text)
    # Remove standalone math expressions: 15*8=120
    text = re.sub(r'\d+\s*[*x]\s*\d+\s*=\s*\d+', '', text)
    # Remove "Toplam" (total) expressions: (3kg Toplam), 3kg toplam, 3kg (Toplam: 6kg)
    text = re.sub(r'\(?[\d.,]+\s*(?:gr|g|kg|ml|lt|l)\s*\(?[Tt]oplam[^)]*\)?', '', text, flags=re.IGNORECASE)
    # Remove size + parenthetical with Toplam: "3kg (Toplam)" "3kg (Toplam: 6kg)"
    text = re.sub(r'[\d.,]+\s*(?:gr|g|kg|ml|lt|l)\s*\([^)]*[Tt]oplam[^)]*\)', '', text, flags=re.IGNORECASE)
    # Remove isolated "toptan", "fiyat", "toplam"
    text = re.sub(r'\b(?:toptan|fiyat|toplam)\b', '', text, flags=re.IGNORECASE)
    # Collapse multiple spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_per_unit_size(description):
    """Extract the per-unit size from a description, preferring individual
    unit sizes over total/aggregate sizes.
    Examples:
        "250 Gram x 8 paket = 2Kg" -> "250 Gram"
        "3kg Toplam"               -> None (total, skip)
        "675 ml"                   -> "675 ml"
    """
    if not description:
        return None
    size_pattern = r'(\d+(?:[.,]\d+)?)\s*(gr|gram|g|kg|kgr|ml|lt|l|litre)\b'
    matches = list(re.finditer(size_pattern, description, re.IGNORECASE))
    if not matches:
        return None
    if len(matches) == 1:
        # If the single match is explicitly a total, skip it
        start = max(0, matches[0].start() - 20)
        end = min(len(description), matches[0].end() + 20)
        context = description[start:end].lower()
        if 'toplam' in context:
            return None
        return matches[0].group(0)
    # Multiple sizes: check for multiplication/total patterns
    has_multiply = bool(re.search(r'\d+\s*(?:gr|g|kg|ml|lt|l)\s*[x*]\s*\d+', description, re.IGNORECASE))
    has_total = bool(re.search(r'[=]\s*\d+\s*(?:gr|g|kg|ml|lt|l)', description, re.IGNORECASE))
    has_toplam = bool(re.search(r'toplam', description, re.IGNORECASE))
    if has_multiply or has_total:
        # Multiplication/equals formula found — return smallest (per-unit) size
        def to_base_unit(value_str, unit_str):
            val = float(value_str.replace(',', '.'))
            u = unit_str.lower()
            if u in ('kg', 'kgr', 'lt', 'l', 'litre'):
                return val * 1000
            return val
        sized = []
        for m in matches:
            base = to_base_unit(m.group(1), m.group(2))
            sized.append((base, m.group(0)))
        sized.sort(key=lambda x: x[0])
        return sized[0][1]  # Return the smallest (per-unit)
    elif has_toplam:
        # "Toplam" label without formula — return size OUTSIDE the Toplam parenthetical
        # e.g. "3kg (Toplam 6kg)" → return "3kg" (outside), skip "6kg" (inside Toplam)
        toplam_paren = re.search(r'\([^)]*[Tt]oplam[^)]*\)', description)
        if toplam_paren:
            non_toplam = [m for m in matches
                          if not (m.start() >= toplam_paren.start() and m.end() <= toplam_paren.end())]
            if non_toplam:
                return non_toplam[0].group(0)
        return None
    # No multiplication pattern — return the first match
    return matches[0].group(0)


def extract_items_per_pkg(description):
    """Extract items-per-package count from description pack patterns.

    Matches Turkish pack-size suffixes like "8li", "12'li", "4'lü", "6'lı".
    These indicate how many individual items are INSIDE each package.

    Examples:
        "Havlu 8li (15 paket 15*8=120 adet)" -> 8
        "Çaykur Altınbaş 500gr"              -> 0 (no pack pattern)
        "250 Gram x 8 paket = 2Kg"           -> 0 (no 'li pattern)
    """
    if not description:
        return 0
    # Match patterns: 8li, 8'li, 12'li, 4'lü, 6'lı, 4'lu (Turkish vowel harmony)
    # Uses double-quoted raw string to avoid apostrophe escaping issues
    match = re.search(r"(\d+)\s*'?l[iıüu]", description, re.IGNORECASE)
    if match:
        items = int(match.group(1))
        # Sanity check: items per package should be 2-100
        if 2 <= items <= 100:
            return items
    return 0


def is_trusted_store(source):
    """Check if a store is in the trusted stores list."""
    if not source:
        return False
    normalized = normalize_turkish(source)
    return any(normalize_turkish(store) in normalized for store in TRUSTED_STORES)


def is_accessory_bundle(title):
    """Check if title contains accessory/bundle indicators."""
    if not title:
        return False
    normalized_title = normalize_turkish(title)

    for keyword in ACCESSORY_KEYWORDS:
        if normalize_turkish(keyword) in normalized_title:
            return True

    # Check for "X + Y" pattern (combo products) e.g. "100gr + 50 bardak"
    if re.search(r'\d+\s*(?:gr|g|ml)\s*\+\s*\d+', title, re.IGNORECASE):
        return True

    return False


def is_relevant_product(title, searched_name, searched_brand):
    """Validate product title actually matches what we're searching for."""
    if not title:
        return False

    normalized_title = normalize_turkish(title)
    normalized_name = normalize_turkish(searched_name)
    normalized_brand = normalize_turkish(searched_brand)

    # Must contain the brand (if specified) — fuzzy match allowed
    if normalized_brand and len(normalized_brand) > 2:
        brand_words = [w for w in normalized_brand.split() if len(w) > 2]
        title_words = normalized_title.split()
        brand_match = any(fuzzy_brand_match(bw, title_words) for bw in brand_words)
        if not brand_match:
            return False

    # Must contain key product words
    for key, keywords in PRODUCT_KEYWORDS.items():
        if normalize_turkish(key) in normalized_name:
            has_keyword = any(normalize_turkish(kw) in normalized_title for kw in keywords)
            if not has_keyword:
                return False
            break

    return True


def extract_quantity_from_title(title, unit, product_name):
    """
    Extract quantity from product title.
    Returns dict with qty, confident, is_bundle, pattern_used.
    """
    if not title:
        return {'qty': 1, 'confident': False}

    normalized_title = normalize_turkish(title)
    original_title = title.lower()
    detected_qty = 1
    confident = False

    # ========== FIRST: Check if this is an accessory bundle ==========
    if is_accessory_bundle(title):
        before_plus = title.split('+')[0]
        main_match = re.search(r'(\d+)\s*(?:gr|g|kg|ml|l)(?:\s|$)', before_plus, re.IGNORECASE)
        if main_match:
            return {'qty': 1, 'confident': True, 'is_bundle': True}
        return {'qty': 1, 'confident': False, 'is_bundle': True}

    # ========== UNIT-SPECIFIC PATTERNS ==========

    # For KG unit
    if unit == 'kg':
        kg_patterns = [
            r'(\d+(?:[.,]\d+)?)\s*kg(?:r|ram)?',
            r'(\d+(?:[.,]\d+)?)\s*kilo',
        ]
        for pattern in kg_patterns:
            match = re.search(pattern, normalized_title, re.IGNORECASE)
            if match:
                detected_qty = float(match.group(1).replace(',', '.'))
                confident = True
                break

        if not confident:
            gram_match = re.search(r'(\d+)\s*gr(?:am)?', normalized_title, re.IGNORECASE)
            if gram_match and 'kg' not in normalized_title:
                detected_qty = int(gram_match.group(1)) / 1000
                confident = True

        return {'qty': detected_qty, 'confident': confident}

    # For LITRE unit
    if unit in ('litre', 'lt'):
        litre_patterns = [
            r'(\d+(?:[.,]\d+)?)\s*l(?:t|itre)?(?!\w)',
            r'(\d+(?:[.,]\d+)?)\s*litre',
        ]
        for pattern in litre_patterns:
            match = re.search(pattern, normalized_title, re.IGNORECASE)
            if match:
                detected_qty = float(match.group(1).replace(',', '.'))
                confident = True
                break

        if not confident:
            ml_match = re.search(r'(\d+)\s*ml', normalized_title, re.IGNORECASE)
            if ml_match:
                detected_qty = int(ml_match.group(1)) / 1000
                confident = True

        return {'qty': detected_qty, 'confident': confident}

    # ========== BULK PACK PATTERNS (High confidence) ==========
    bulk_patterns = [
        # "x 50 Adet" or "X50 Adet"
        {'pattern': r'[x\u00d7]\s*(\d+)\s*adet', 'name': 'X N adet', 'multiply': False},
        # "50 Adet" at start or after size
        {'pattern': r'(?:^|\d+\s*(?:gr|g|ml|kg)\s*)(\d+)\s*adet', 'name': 'N adet after size', 'multiply': False},
        # "50'li Set" or "50'li Paket"
        {'pattern': r"(\d+)\s*['\u00b4\u2018\u2019`]?\s*l[iiuu]\s*(?:set|paket|koli)", 'name': "N'li set/paket", 'multiply': False},
        # "(50 Adet)" or "(50'li)" in parentheses
        {'pattern': r"\((\d+)\s*(?:adet|'?l[iiuu])\)", 'name': 'parentheses count', 'multiply': False},
        # "100gr x 50" - size times count
        {'pattern': r'\d+\s*(?:gr|g|ml)\s*[x\u00d7]\s*(\d+)(?:\s|$|adet)', 'name': 'size x count', 'multiply': False},
        # "2 Koli*25" or "2 Koli x 25"
        {'pattern': r'(\d+)\s*koli\s*[x\u00d7\*]\s*(\d+)', 'name': 'koli pattern', 'multiply': True},
    ]

    for bp in bulk_patterns:
        match = re.search(bp['pattern'], original_title, re.IGNORECASE)
        if not match:
            match = re.search(bp['pattern'], normalized_title, re.IGNORECASE)
        if match:
            if bp['multiply'] and match.lastindex and match.lastindex >= 2:
                detected_qty = int(match.group(1)) * int(match.group(2))
            else:
                detected_qty = int(match.group(1))

            if 2 <= detected_qty <= 500:
                return {'qty': detected_qty, 'confident': True, 'pattern_used': bp['name']}

    # ========== STANDARD PACK PATTERNS (Medium confidence) ==========
    pack_patterns = [
        # Turkish: 10'lu, 8'li
        {'pattern': r"(\d+)\s*['\u00b4\u2018\u2019`]\s*l[iiuu]", 'name': 'Turkish apostrophe'},
        # Turkish without apostrophe: 10lu, 8li (but NOT after + sign)
        {'pattern': r'(?<!\+)(?<!\+ )(\d+)\s*l[iiuu](?:\s|$|,|\)|paket|koli|set|kutu)', 'name': 'Turkish no apostrophe'},
        # "N adet" standalone
        {'pattern': r'(?:^|\s)(\d+)\s*adet(?:\s|$|,|\))', 'name': 'N adet'},
        # Tablet/Capsule
        {'pattern': r'(\d+)\s*(?:tablet|kapsul|kapsul)', 'name': 'Tablet'},
        # Rulo
        {'pattern': r'(\d+)\s*(?:rulo|roll)', 'name': 'Rulo'},
    ]

    # Non-count unit words: when a number is followed by these, it's a measurement NOT a pack count
    # "200 yaprak" = 200 sheets (per roll), "500 ml" = volume, "250 gr" = weight
    NON_COUNT_UNITS = {'yaprak', 'sayfa', 'ml', 'gr', 'gram', 'kg', 'lt', 'litre', 'cm', 'mm', 'm', 'metre'}

    for pp in pack_patterns:
        match = re.search(pp['pattern'], original_title, re.IGNORECASE)
        if not match:
            match = re.search(pp['pattern'], normalized_title, re.IGNORECASE)
        if match:
            qty = int(match.group(1))
            if 2 <= qty <= 500:
                # Check if this number is actually a measurement, not a pack count
                # Look at the broader context around the match
                match_end = match.end()
                search_text = original_title if match.string == original_title else normalized_title
                after_match = search_text[match_end:match_end + 15].strip().lower()
                # Also check what comes after the number itself (before the pattern suffix)
                num_str = str(qty)
                num_pos = search_text.lower().find(num_str)
                if num_pos >= 0:
                    after_num = search_text[num_pos + len(num_str):num_pos + len(num_str) + 12].strip().lower()
                    first_word_after = after_num.split()[0] if after_num.split() else ''
                    if normalize_turkish(first_word_after) in NON_COUNT_UNITS:
                        # This is a measurement (e.g., "200 yaprak"), not a pack count — skip
                        continue
                return {'qty': qty, 'confident': True, 'pattern_used': pp['name']}

    # No quantity pattern found → this is a single item.
    # Treat as confident qty=1 so single items compete fairly with
    # multi-packs in per-unit-price sorting (prevents over-purchase bias).
    return {'qty': 1, 'confident': True, 'pattern_used': 'single_item_default'}


def calculate_relevance_score(title, source, original_name, original_brand):
    """Calculate relevance score for a search result."""
    if not title:
        return 0

    normalized_title = normalize_turkish(title)
    normalized_name = normalize_turkish(original_name)
    normalized_brand = normalize_turkish(original_brand)

    score = 100

    # PENALTY: Accessory bundles
    if is_accessory_bundle(title):
        score -= 40

    # Check if product name matches
    name_words = [w for w in normalized_name.split() if len(w) > 2]
    name_match_count = sum(1 for w in name_words if w in normalized_title)

    if name_words:
        name_match_ratio = name_match_count / len(name_words)
        if name_match_ratio < 0.5:
            score -= 50
        elif name_match_ratio < 1:
            score -= 20

    # Check if brand matches (fuzzy)
    if normalized_brand and len(normalized_brand) > 2:
        brand_words = [w for w in normalized_brand.split() if len(w) > 2]
        title_words = normalized_title.split()
        brand_match = any(fuzzy_brand_match(bw, title_words) for bw in brand_words)
        if not brand_match:
            score -= 35

    # Penalty for wrong product categories
    for cat in WRONG_CATEGORIES:
        normalized_cat = normalize_turkish(cat)
        if normalized_cat in normalized_title and normalize_turkish(cat.split()[0]) not in normalized_name:
            score -= 60

    # BONUS: Bulk indicators
    bulk_indicators = ['toptan', 'koli', 'set', 'paket', 'x 50', 'x50', '50 adet']
    for indicator in bulk_indicators:
        if normalize_turkish(indicator) in normalized_title:
            score += 25
            break

    # Bonus for trusted stores
    if is_trusted_store(source):
        score += 20

    return max(0, score)


def build_search_query(product_name, brand, description, unit, quantity, include_quantity=True):
    """
    Build an optimized search query for Google Shopping.
    No longer appends 'toptan' or 'fiyat' — these biased results toward
    bulk industrial listings with inflated prices.
    """
    name = (product_name or '').strip()
    brand = (brand or '').strip()
    description = (description or '').strip()
    unit = (unit or 'adet').lower().strip()
    quantity = quantity or 1

    name_lower = normalize_turkish(name)

    # Check if this is a generic product that needs description-based search
    use_description = False
    matched_generic = None
    for generic in GENERIC_PRODUCT_NAMES:
        if normalize_turkish(generic) in name_lower and description:
            use_description = True
            matched_generic = generic
            break

    if use_description:
        cleaned_desc = clean_description(description)
        # Use mapped search term if available (e.g., "Rulo Peçete" → "kağıt havlu")
        # This uses Google Shopping's indexed terms instead of colloquial names
        mapped_name = GENERIC_NAME_MAP.get(matched_generic, name)
        # Remove words from description that are already in the mapped name
        # e.g., mapped "kağıt havlu" + desc "Havlu 8li" → "kağıt havlu 8li" (no dup)
        mapped_words = set(normalize_turkish(mapped_name).split())
        desc_parts = [w for w in cleaned_desc.split()
                      if normalize_turkish(w) not in mapped_words]
        unique_desc = ' '.join(desc_parts)
        search_query = f"{mapped_name} {unique_desc} {brand}".strip()
    else:
        search_query = f"{name} {brand}".strip()

    # Add unit hint (not 'adet') for non-count units like kg, paket, litre
    if include_quantity and unit and unit not in ('adet',):
        search_query += f" {unit}"

    # Extract per-unit size first (needed to decide on quantity hint)
    per_unit_size = None
    if description:
        per_unit_size = extract_per_unit_size(description)

    # Check if description already has a pack-size pattern like "8li", "4lü", "12'li"
    # If so, don't add order quantity as 'li — description already specifies pack size
    # e.g. "Havlu 8li" — "8li" IS the pack size, qty=15 is order count (NOT a pack size)
    desc_has_pack_hint = bool(re.search(r"\d+['\u2019]?l[i\u0131\u00fc\u00fcu]", description or '', re.IGNORECASE))

    # Add pack-size hints based on quantity
    if include_quantity and quantity:
        qty_int = int(quantity)
        if desc_has_pack_hint:
            # Description already specifies pack size (e.g., "8li") — skip order qty hint
            pass
        elif 6 <= qty_int <= 24:
            # Large multi-packs: always add quantity hint
            search_query += f" {qty_int}'li"
        elif 2 <= qty_int <= 5 and not per_unit_size:
            # Small quantities: add qty hint only when no size info found
            # e.g. Islak Mendil qty=4 (no size) → "4'li"
            # e.g. Pril qty=4 (675 ml) → no hint (size already in query)
            # e.g. Peros qty=2 (3kg) → no hint (size already in query)
            search_query += f" {qty_int}'li"

    # Append per-unit size if found and not already in query
    if per_unit_size and per_unit_size.lower() not in search_query.lower():
        search_query += f" {per_unit_size}"

    return search_query.strip()


# ============================================
# ODOO MODELS
# ============================================

class TotalinePriceSearch(models.Model):
    _name = 'totaline.price.search'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Totaline Price Search'
    _order = 'create_date desc'

    name = fields.Char(string='Search Name', required=True, default=lambda self: f"Search {fields.Datetime.now()}")
    state = fields.Selection([
        ('draft', 'Draft'),
        ('searching', 'Searching'),
        ('done', 'Completed'),
        ('error', 'Error'),
    ], string='Status', default='draft')

    # Excel Upload
    excel_file = fields.Binary(string='Excel File', attachment=True)
    excel_filename = fields.Char(string='Filename')
    excel_sheet_name = fields.Char(string='Sheet Name',
                                    help='Select which sheet to import from')
    excel_sheet_names = fields.Char(string='Available Sheets', readonly=True,
                                     help='Comma-separated list of sheet names')

    @api.onchange('excel_sheet_name')
    def _onchange_excel_sheet_name(self):
        """Re-parse Excel when a different sheet is selected."""
        if self.excel_file and self.excel_sheet_name:
            self._parse_excel_sheet(self.excel_sheet_name)

    @api.onchange('excel_file')
    def _onchange_excel_file(self):
        """Parse Excel file when uploaded and populate product lines"""
        if not self.excel_file:
            self.excel_sheet_names = False
            self.excel_sheet_name = False
            return

        if not OPENPYXL_AVAILABLE:
            raise UserError('Excel parsing library (openpyxl) is not installed.')

        try:
            file_content = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()

            if len(sheet_names) > 1:
                # Multiple sheets: show selection and wait
                self.excel_sheet_names = ', '.join(sheet_names)
                self.excel_sheet_name = sheet_names[0]  # Default to first
                self._parse_excel_sheet(sheet_names[0])
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': 'Multiple Sheets Detected',
                        'message': f'Found {len(sheet_names)} sheets: {", ".join(sheet_names)}. '
                                   f'Loaded "{sheet_names[0]}". Change the Sheet Name field to switch.',
                        'type': 'warning',
                        'sticky': True,
                    }
                }
            else:
                self.excel_sheet_names = False
                self.excel_sheet_name = False
                self._parse_excel_sheet(sheet_names[0] if sheet_names else None)

        except UserError:
            raise
        except Exception as e:
            _logger.exception('Failed to parse Excel file')
            raise UserError(f'Failed to parse Excel file: {str(e)}')

    def _parse_excel_sheet(self, sheet_name=None):
        """Parse a specific sheet from the uploaded Excel file."""
        if not self.excel_file:
            return

        try:
            file_content = base64.b64decode(self.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))

            if sheet_name and sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Clear existing lines
            self.search_line_ids = [(5, 0, 0)]

            # Find header row and map columns
            headers = {}
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    header_name = str(cell.value).strip().lower()
                    headers[header_name] = col_idx

            # Map Turkish/English column names
            column_mapping = {
                'product_name': ['\u00fcr\u00fcn ad\u0131', 'urun adi', 'product name', 'product', '\u00fcr\u00fcn', 'urun'],
                'brand': ['marka', 'brand'],
                'quantity': ['miktar', 'quantity', 'qty', 'adet'],
                'unit': ['birim', 'unit'],
                'description': ['a\u00e7\u0131klama', 'aciklama', 'description', 'desc', 'detay'],
                'package_quantity': ['package qty', 'package quantity', 'paket adedi', 'paket miktar'],
                'items_per_package': ['items per package', 'items/pkg', 'paket icerigi', 'paket ici adet'],
            }

            # Find column indices
            col_indices = {}
            for field_name, names in column_mapping.items():
                for col_name in names:
                    if col_name in headers:
                        col_indices[field_name] = headers[col_name]
                        break

            if 'product_name' not in col_indices:
                raise UserError('Could not find product name column. Expected: "\u00dcr\u00fcn Ad\u0131" or "Product Name"')

            # Parse data rows
            lines = []
            for row_idx in range(2, sheet.max_row + 1):
                product_name = sheet.cell(row=row_idx, column=col_indices.get('product_name', 1)).value

                if not product_name:
                    continue  # Skip empty rows

                line_data = {
                    'product_name': str(product_name).strip(),
                    'brand': '',
                    'quantity': 1,
                    'unit': '',
                    'description': '',
                }

                if 'brand' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['brand']).value
                    line_data['brand'] = str(val).strip() if val else ''

                if 'quantity' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['quantity']).value
                    try:
                        line_data['quantity'] = float(val) if val else 1
                    except (ValueError, TypeError):
                        line_data['quantity'] = 1

                if 'unit' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['unit']).value
                    line_data['unit'] = str(val).strip() if val else ''

                if 'description' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['description']).value
                    line_data['description'] = str(val).strip() if val else ''

                if 'package_quantity' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['package_quantity']).value
                    try:
                        line_data['package_quantity'] = float(val) if val else 1
                    except (ValueError, TypeError):
                        line_data['package_quantity'] = 1

                if 'items_per_package' in col_indices:
                    val = sheet.cell(row=row_idx, column=col_indices['items_per_package']).value
                    try:
                        line_data['items_per_package'] = float(val) if val else 1
                    except (ValueError, TypeError):
                        line_data['items_per_package'] = 1

                # Sync: if package fields were provided, compute quantity from them
                pkg_qty = line_data.get('package_quantity', 1)
                items_pkg = line_data.get('items_per_package', 1)
                if pkg_qty > 1 or items_pkg > 1:
                    line_data['quantity'] = pkg_qty * items_pkg
                elif line_data.get('quantity', 1) > 1 and 'package_quantity' not in col_indices:
                    # Excel only has Miktar (quantity), no package columns:
                    # put quantity into package_quantity so it shows in the UI
                    line_data['package_quantity'] = line_data['quantity']

                lines.append((0, 0, line_data))

            self.search_line_ids = lines
            _logger.info(f'Parsed {len(lines)} products from sheet "{sheet_name or "active"}"')

        except UserError:
            raise
        except Exception as e:
            _logger.exception('Failed to parse Excel sheet')
            raise UserError(f'Failed to parse Excel file: {str(e)}')

    # Results
    search_line_ids = fields.One2many('totaline.price.search.line', 'search_id', string='Search Results')

    # Totals
    total_best_price = fields.Float(string='Total (Best Price)', compute='_compute_totals', store=True)
    total_second_price = fields.Float(string='Total (2nd Price)', compute='_compute_totals', store=True)
    total_third_price = fields.Float(string='Total (3rd Price)', compute='_compute_totals', store=True)

    # Info
    product_count = fields.Integer(string='Product Count', compute='_compute_product_count')
    search_date = fields.Datetime(string='Search Date', readonly=True)
    error_message = fields.Text(string='Error Message')

    # Scheduling fields
    is_scheduled = fields.Boolean(string='Scheduled Search', default=False,
                                   help='Enable automatic scheduled price searches')
    schedule_frequency = fields.Selection([
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
    ], string='Frequency', default='weekly')
    schedule_time = fields.Float(string='Search Time', default=8.0,
                                  help='Time of day to run search (24h format, e.g. 8.0 = 08:00)')
    schedule_day = fields.Selection([
        ('0', 'Monday'),
        ('1', 'Tuesday'),
        ('2', 'Wednesday'),
        ('3', 'Thursday'),
        ('4', 'Friday'),
        ('5', 'Saturday'),
        ('6', 'Sunday'),
    ], string='Day of Week', default='0', help='For weekly schedules')
    schedule_day_of_month = fields.Integer(string='Day of Month', default=1,
                                            help='For monthly schedules (1-28)')
    next_scheduled_date = fields.Datetime(string='Next Scheduled Run', compute='_compute_next_scheduled', store=True)
    last_scheduled_run = fields.Datetime(string='Last Scheduled Run', readonly=True)

    @api.depends('is_scheduled', 'schedule_frequency', 'schedule_time', 'schedule_day', 'schedule_day_of_month')
    def _compute_next_scheduled(self):
        from datetime import datetime, timedelta
        for record in self:
            if not record.is_scheduled:
                record.next_scheduled_date = False
                continue

            now = fields.Datetime.now()
            hour = int(record.schedule_time)
            minute = int((record.schedule_time % 1) * 60)

            if record.schedule_frequency == 'daily':
                next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if next_run <= now:
                    next_run += timedelta(days=1)
            elif record.schedule_frequency == 'weekly':
                target_day = int(record.schedule_day or '0')
                days_ahead = target_day - now.weekday()
                if days_ahead < 0:
                    days_ahead += 7
                next_run = now + timedelta(days=days_ahead)
                next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
                if next_run <= now:
                    next_run += timedelta(days=7)
            else:  # monthly
                day = min(record.schedule_day_of_month or 1, 28)
                next_run = now.replace(day=day, hour=hour, minute=minute, second=0, microsecond=0)
                if next_run <= now:
                    if now.month == 12:
                        next_run = next_run.replace(year=now.year + 1, month=1)
                    else:
                        next_run = next_run.replace(month=now.month + 1)

            record.next_scheduled_date = next_run

    @api.depends('search_line_ids')
    def _compute_product_count(self):
        for record in self:
            record.product_count = len(record.search_line_ids)

    @api.depends('search_line_ids.total_cost_1', 'search_line_ids.total_cost_2', 'search_line_ids.total_cost_3')
    def _compute_totals(self):
        for record in self:
            record.total_best_price = sum(line.total_cost_1 or 0 for line in record.search_line_ids)
            record.total_second_price = sum(line.total_cost_2 or 0 for line in record.search_line_ids)
            record.total_third_price = sum(line.total_cost_3 or 0 for line in record.search_line_ids)

    # ============================================
    # SERPAPI DIRECT INTEGRATION
    # ============================================

    def _get_serpapi_key(self):
        """Get SerpAPI key from Odoo System Parameters."""
        api_key = self.env['ir.config_parameter'].sudo().get_param('totaline.serpapi_key', '')
        if not api_key:
            raise UserError(
                'SerpAPI key not configured.\n\n'
                'Go to Settings > Technical > Parameters > System Parameters\n'
                'and create a parameter:\n'
                '  Key: totaline.serpapi_key\n'
                '  Value: your SerpAPI key'
            )
        return api_key

    def _search_serpapi(self, search_query, api_key):
        """Call SerpAPI Google Shopping directly and return results."""
        params = {
            'engine': 'google_shopping',
            'q': search_query,
            'location': 'Turkey',
            'gl': 'tr',
            'hl': 'tr',
            'api_key': api_key,
        }

        try:
            response = requests.get(
                'https://serpapi.com/search.json',
                params=params,
                timeout=30,
            )

            if response.status_code == 200:
                return response.json()
            elif response.status_code == 429:
                _logger.error('SerpAPI rate limit exceeded (429)')
                raise UserError(
                    'SerpAPI rate limit exceeded (429 Too Many Requests).\n\n'
                    'Please wait a few minutes and try again.\n'
                    'If you are on a free plan, check your daily/monthly quota:\n'
                    'https://serpapi.com/manage-api-key'
                )
            elif response.status_code == 401:
                _logger.error('SerpAPI authentication failed (401)')
                raise UserError(
                    'SerpAPI API key is invalid (401 Unauthorized).\n\n'
                    'Please check your API key:\n'
                    'Settings → Technical → System Parameters → totaline.serpapi_key'
                )
            elif response.status_code == 403:
                _logger.error('SerpAPI quota exhausted (403)')
                raise UserError(
                    'SerpAPI search quota exhausted (403 Forbidden).\n\n'
                    'Your free search credits may have run out.\n'
                    'Enter a new API key or upgrade your plan:\n'
                    'https://serpapi.com/manage-api-key\n\n'
                    'To change the key:\n'
                    'Settings → Technical → System Parameters → totaline.serpapi_key'
                )
            else:
                _logger.warning(f'SerpAPI returned status {response.status_code}: {response.text[:200]}')
                return {}

        except UserError:
            raise
        except requests.exceptions.Timeout:
            _logger.warning(f'SerpAPI timeout for query: {search_query}')
            return {}
        except Exception as e:
            _logger.exception(f'SerpAPI request failed: {e}')
            return {}

    def _search_serpapi_immersive(self, page_token, api_key):
        """
        Call SerpAPI google_immersive_product engine to get verified store offers.

        This is Step 2 of the two-step search:
        Step 1: google_shopping → find the best matching product (with page_token)
        Step 2: google_immersive_product → get verified multi-seller offers with
                real prices, shipping costs, and direct store links.

        Returns raw JSON response or empty dict on failure.
        """
        params = {
            'engine': 'google_immersive_product',
            'page_token': page_token,
            'api_key': api_key,
        }

        try:
            response = requests.get(
                'https://serpapi.com/search.json',
                params=params,
                timeout=30,
            )

            if response.status_code == 200:
                data = response.json()
                product_info = data.get('product_results', {})
                # Sellers can be under product_results.stores (Turkish) or sellers_results.online_sellers
                stores = product_info.get('stores', [])
                sellers_online = data.get('sellers_results', {}).get('online_sellers', []) if not stores else []
                seller_count = len(stores or sellers_online)
                _logger.info(
                    f'Immersive product: {seller_count} offers for '
                    f'"{product_info.get("title", "unknown")}"'
                )
                return data
            else:
                _logger.warning(f'SerpAPI immersive product returned status {response.status_code}')
                return {}

        except Exception as e:
            _logger.warning(f'SerpAPI immersive product request failed: {e}')
            return {}

    def _format_immersive_results(self, immersive_data, product_name, brand, quantity, unit, shopping_detected_qty=1):
        """
        Format immersive product store offers into the standard price result dict.

        The immersive product API returns verified offers from multiple stores
        for ONE specific product. Each store has: price, shipping, total, direct link.

        Args:
            immersive_data: Raw response from google_immersive_product
            product_name: Original product name from Excel
            brand: Original brand from Excel
            quantity: Order quantity needed
            unit: Unit type (adet, paket, etc.)
            shopping_detected_qty: Quantity detected from the shopping result title
                                   (the product that led to this immersive lookup)

        Returns:
            Formatted result dict (same structure as _format_search_results)
            or None if no valid offers found.
        """
        product_info = immersive_data.get('product_results', {})

        # Extract sellers from multiple possible response paths
        # SerpAPI structures this differently by locale/product type
        sellers = []

        # Path 1: product_results.stores (Turkish locale — confirmed working)
        sellers = product_info.get('stores', [])

        # Path 2: sellers_results.online_sellers (English locale / other product types)
        if not sellers:
            sellers_results = immersive_data.get('sellers_results', {})
            if isinstance(sellers_results, dict):
                sellers = sellers_results.get('online_sellers', [])

        # Path 3: Direct online_sellers at root level
        if not sellers:
            sellers = immersive_data.get('online_sellers', [])

        # Path 4: Other possible keys inside product_results
        if not sellers:
            sellers = product_info.get('sellers', []) or product_info.get('online_sellers', [])

        if not sellers:
            _logger.info(f'No sellers found in immersive product for "{product_name}"')
            return None

        original_qty = int(quantity or 1)
        original_unit = (unit or 'adet').lower()
        verified_title = product_info.get('title', '')
        verified_brand = product_info.get('brand', '')

        # Process each seller offer
        processed = []
        for seller in sellers:
            # Extract price — try multiple field names
            price = 0
            try:
                price = float(
                    seller.get('extracted_price', 0) or
                    seller.get('base_price_extracted', 0) or
                    0
                )
            except (ValueError, TypeError):
                continue

            if price <= 0:
                continue

            # Extract shipping cost
            shipping = 0
            try:
                shipping = float(seller.get('shipping_extracted', 0) or 0)
            except (ValueError, TypeError):
                shipping = 0

            # Total price (price + shipping)
            total = 0
            try:
                total = float(seller.get('extracted_total', 0) or 0)
            except (ValueError, TypeError):
                total = 0
            if total <= 0:
                total = price + shipping

            store_name = seller.get('name', seller.get('source', 'Unknown'))
            link = seller.get('link', '')
            tag = seller.get('tag', '')
            trusted = is_trusted_store(store_name)

            processed.append({
                'source': store_name,
                'link': link,
                'price': price,
                'shipping': shipping,
                'total': total,
                'tag': tag,
                'trusted': trusted,
            })

        if not processed:
            _logger.info(f'No valid seller prices in immersive product for "{product_name}"')
            return None

        # Sort: trusted stores first, then by total price (including shipping)
        processed.sort(key=lambda s: (not s['trusted'], s['total']))

        top_3 = processed[:3]

        # Use the detected_qty from the shopping result that led us here
        detected_qty = shopping_detected_qty or 1

        # Calculate cost to fulfill for the warning
        if detected_qty >= 1 and original_qty > 0:
            packs_needed = math.ceil(original_qty / detected_qty)
        else:
            packs_needed = original_qty

        # Build result dict (same structure as _format_search_results)
        result = {
            'price_1': top_3[0]['price'] if len(top_3) > 0 else 0,
            'store_1': top_3[0]['source'] if len(top_3) > 0 else '',
            'url_1': top_3[0]['link'] if len(top_3) > 0 else '',
            'price_2': top_3[1]['price'] if len(top_3) > 1 else 0,
            'store_2': top_3[1]['source'] if len(top_3) > 1 else '',
            'url_2': top_3[1]['link'] if len(top_3) > 1 else '',
            'price_3': top_3[2]['price'] if len(top_3) > 2 else 0,
            'store_3': top_3[2]['source'] if len(top_3) > 2 else '',
            'url_3': top_3[2]['link'] if len(top_3) > 2 else '',
            # All stores sell the SAME product → same detected quantity
            'detected_qty_1': detected_qty,
            'detected_qty_2': detected_qty,
            'detected_qty_3': detected_qty,
            'warning': '',
        }

        # Build informative warning with verification + shipping info
        warnings = []

        # Show verified product title (proves we found the right product)
        if verified_title:
            short_title = verified_title[:60] + ('...' if len(verified_title) > 60 else '')
            warnings.append(f"✓ {short_title}")

        # Show shipping cost for best offer
        if top_3[0].get('shipping', 0) > 0:
            warnings.append(f"+₺{top_3[0]['shipping']:.0f} kargo")

        # Show tag if present (e.g., "En iyi fiyat", "En popüler")
        if top_3[0].get('tag'):
            warnings.append(top_3[0]['tag'])

        # Show pack info if multi-pack
        if detected_qty > 1 and original_qty > 1:
            warnings.append(f"{detected_qty}'li paket → {packs_needed}x sipariş")

        result['warning'] = ' | '.join(warnings) if warnings else ''

        best_shipping = top_3[0].get('shipping', 0)
        shipping_info = f' (+₺{best_shipping:.0f} kargo)' if best_shipping > 0 else ''
        _logger.info(
            f'Immersive results for "{product_name}": '
            f'{len(processed)} sellers, best: ₺{result["price_1"]:.0f} at {result["store_1"]}{shipping_info}'
        )

        return result

    def _format_search_results(self, serp_results, product_name, brand, quantity, unit, description, relevance_threshold=40):
        """
        Process SerpAPI results and return top 3 price results.

        Returns dict with: price_1, store_1, url_1, price_2, ..., price_3, ..., warning
        """
        original_name = product_name or ''
        original_brand = brand or ''
        original_quantity = int(quantity or 1)
        original_unit = (unit or 'adet').lower()

        shopping_results = serp_results.get('shopping_results', [])
        processed_results = []

        for item in shopping_results:
            title = item.get('title', '')
            price = 0
            try:
                price = float(item.get('extracted_price', 0) or 0)
            except (ValueError, TypeError):
                continue

            source = item.get('source', 'Unknown')
            # SerpAPI Google Shopping: 'link' is usually empty,
            # 'product_link' has the Google Shopping product page URL
            link = item.get('product_link', '') or item.get('link', '')
            # Immersive product page token for two-step verified search
            page_token = item.get('immersive_product_page_token', '')

            if price <= 0:
                continue

            # Check relevance
            if not is_relevant_product(title, original_name, original_brand):
                continue

            # Check if accessory bundle
            is_bundle_product = is_accessory_bundle(title)

            # Extract quantity from title
            quantity_info = extract_quantity_from_title(title, original_unit, original_name)
            detected_qty = quantity_info.get('qty', 1)
            confident = quantity_info.get('confident', False)

            # Calculate per-unit price for sorting
            # IMPORTANT: We always display the RAW Google Shopping listing price (not multiplied).
            # Quantity multiplication only happens when creating Purchase Orders.
            warning = None
            per_unit_price = price  # price per single unit for sorting

            if is_bundle_product:
                per_unit_price = price
                warning = f"\u2139\ufe0f Gift set / bundle product"
            elif confident and detected_qty > 1:
                per_unit_price = price / detected_qty
                if original_quantity > 1:
                    packs_needed = math.ceil(original_quantity / detected_qty)
                    warning = f"\u2139\ufe0f Pack of {detected_qty} — need {packs_needed} packs for {original_quantity} {original_unit}"
                else:
                    warning = f"\u2139\ufe0f Package has {detected_qty} {original_unit}"
            elif not confident and original_quantity > 1:
                warning = f"\u26a0\ufe0f {original_quantity} {original_unit} needed — check listing qty"

            # total_price = raw listing price (NO multiplication)
            total_price = price
            relevance_score = calculate_relevance_score(title, source, original_name, original_brand)
            trusted = is_trusted_store(source)

            # Exact quantity match bonus
            if confident and detected_qty == original_quantity:
                relevance_score += 30

            # Over-purchase penalty: when detected_qty is way higher than needed,
            # it's likely a misdetection (e.g., "200 yaprak" → qty=200 for 15 packs)
            # or an enormous bulk pack that wastes money.
            qty_misdetected = False
            if detected_qty > 1 and original_quantity > 1:
                ratio = detected_qty / original_quantity
                if ratio > 5:
                    # Extreme mismatch: likely quantity misdetection
                    relevance_score -= 40
                    per_unit_price = price  # Treat as single unit (reset per_unit calc)
                    detected_qty = 1  # Reset to single for cost calculation
                    confident = False  # Mark as not confident
                    qty_misdetected = True
                    warning = f"⚠️ Qty detection unreliable — listing may be single unit"
                elif ratio > 2:
                    # Moderate over-purchase: penalize proportionally
                    relevance_score -= 15

            # Calculate cost_to_fulfill: actual cost to buy enough packs for the order
            # This helps prefer singles over over-sized packs when total cost matters
            # e.g., 4 singles at ₺60 = ₺240 is cheaper than 1x 7-pack at ₺298
            if confident and detected_qty >= 1 and original_quantity > 0:
                packs_needed = math.ceil(original_quantity / detected_qty)
                cost_to_fulfill = price * packs_needed
            else:
                cost_to_fulfill = price * original_quantity  # Assume single unit

            processed_results.append({
                'title': title,
                'source': source,
                'link': link,
                'original_price': price,
                'per_unit_price': per_unit_price,
                'cost_to_fulfill': cost_to_fulfill,
                'detected_qty': detected_qty,
                'confident': confident,
                'is_bundle': is_bundle_product,
                'multiplier': 1,
                'total_price': total_price,
                'warning': warning,
                'relevance_score': relevance_score,
                'trusted': trusted,
                'page_token': page_token,
            })

        # ============================================
        # SELECT TOP 3 RESULTS
        # ============================================

        # Filter trusted results with good relevance
        trusted_results = [
            r for r in processed_results
            if r['trusted'] and r['relevance_score'] >= relevance_threshold
        ]
        # Sort: non-bundles first → confident first → cheapest cost to fulfill order
        # All results already passed relevance threshold + brand matching = correct products.
        # cost_to_fulfill penalizes over-sized packs naturally (e.g., 7-pack for 4 units).
        trusted_results.sort(key=lambda r: (
            r['is_bundle'],           # False < True, so non-bundles first
            not r['confident'],       # confident first
            r['cost_to_fulfill'],     # cheapest total cost to fulfill order
        ))

        untrusted_results = [
            r for r in processed_results
            if not r['trusted'] and r['relevance_score'] >= relevance_threshold
        ]
        untrusted_results.sort(key=lambda r: r['cost_to_fulfill'])

        # Combine: trusted first, then untrusted
        all_sorted = trusted_results + untrusted_results
        top_3 = all_sorted[:3]

        # Build result dict
        result = {
            'price_1': top_3[0]['total_price'] if len(top_3) > 0 else 0,
            'store_1': top_3[0]['source'] if len(top_3) > 0 else '',
            'url_1': top_3[0]['link'] if len(top_3) > 0 else '',
            'price_2': top_3[1]['total_price'] if len(top_3) > 1 else 0,
            'store_2': top_3[1]['source'] if len(top_3) > 1 else '',
            'url_2': top_3[1]['link'] if len(top_3) > 1 else '',
            'price_3': top_3[2]['total_price'] if len(top_3) > 2 else 0,
            'store_3': top_3[2]['source'] if len(top_3) > 2 else '',
            'url_3': top_3[2]['link'] if len(top_3) > 2 else '',
            'detected_qty_1': top_3[0]['detected_qty'] if len(top_3) > 0 else 0,
            'detected_qty_2': top_3[1]['detected_qty'] if len(top_3) > 1 else 0,
            'detected_qty_3': top_3[2]['detected_qty'] if len(top_3) > 2 else 0,
            'page_token_1': top_3[0].get('page_token', '') if len(top_3) > 0 else '',
            'best_title': top_3[0].get('title', '') if len(top_3) > 0 else '',
            'warning': '',
        }

        # Collect warnings from top results
        warnings = []
        for r in top_3:
            if r.get('warning'):
                warnings.append(r['warning'])
                break  # Only first warning

        if not top_3:
            warnings.append('No results found')

        result['warning'] = warnings[0] if warnings else ''

        _logger.info(
            f'Product "{original_name}": found {len(shopping_results)} raw, '
            f'{len(processed_results)} processed, {len(trusted_results)} trusted, '
            f'top price: {result["price_1"]}'
        )

        return result

    # ============================================
    # MAIN SEARCH ACTION
    # ============================================

    def action_search_prices(self):
        """Search prices directly via SerpAPI (no n8n dependency)."""
        self.ensure_one()

        if not self.search_line_ids:
            self.state = 'error'
            self.error_message = 'No products to search. Please add products first.'
            return

        self.state = 'searching'
        self.search_date = fields.Datetime.now()
        self.error_message = False

        try:
            api_key = self._get_serpapi_key()

            success_count = 0
            error_count = 0

            for line in self.search_line_ids:
                try:
                    # ================================================
                    # STEP 1: Google Shopping search (product discovery)
                    # ================================================
                    search_query = build_search_query(
                        product_name=line.product_name,
                        brand=line.brand,
                        description=line.description,
                        unit=line.unit,
                        quantity=line.quantity,
                        include_quantity=True,
                    )

                    _logger.info(f'[Step 1] Shopping search: "{search_query}" for "{line.product_name}"')

                    serp_results = self._search_serpapi(search_query, api_key)

                    # Format shopping results (identifies best matching product)
                    formatted = self._format_search_results(
                        serp_results=serp_results,
                        product_name=line.product_name,
                        brand=line.brand,
                        quantity=line.quantity,
                        unit=line.unit,
                        description=line.description,
                        relevance_threshold=40,
                    )

                    # FALLBACK TIER 1: Relax relevance threshold (no extra API call)
                    if not formatted.get('price_1'):
                        _logger.info(f'No results at threshold=40 for "{line.product_name}", relaxing to 20')
                        formatted = self._format_search_results(
                            serp_results=serp_results,
                            product_name=line.product_name,
                            brand=line.brand,
                            quantity=line.quantity,
                            unit=line.unit,
                            description=line.description,
                            relevance_threshold=20,
                        )

                    # FALLBACK TIER 2: Simplified query (1 extra API call)
                    if not formatted.get('price_1'):
                        fallback_query = f"{line.product_name} {line.brand}".strip()
                        _logger.info(f'Fallback search: "{fallback_query}" for "{line.product_name}"')
                        serp_fallback = self._search_serpapi(fallback_query, api_key)
                        formatted = self._format_search_results(
                            serp_results=serp_fallback,
                            product_name=line.product_name,
                            brand=line.brand,
                            quantity=line.quantity,
                            unit=line.unit,
                            description=line.description,
                            relevance_threshold=20,
                        )
                        if formatted.get('price_1'):
                            formatted['warning'] = (formatted.get('warning', '') + ' (fallback query)').strip()

                    # ================================================
                    # STEP 2: Immersive Product (verified store offers)
                    # ================================================
                    # If Step 1 found a good match with a page_token,
                    # fetch verified store offers with real prices + shipping + direct links.
                    # This costs 1 extra API call but gives much more reliable data.
                    page_token = formatted.get('page_token_1', '')
                    if page_token and formatted.get('price_1'):
                        _logger.info(
                            f'[Step 2] Immersive product lookup for "{line.product_name}" '
                            f'(matched: "{formatted.get("best_title", "")[:50]}")'
                        )

                        immersive_data = self._search_serpapi_immersive(page_token, api_key)

                        if immersive_data:
                            # Use the detected quantity from Step 1's best match
                            shopping_qty = formatted.get('detected_qty_1', 1) or 1

                            immersive_formatted = self._format_immersive_results(
                                immersive_data=immersive_data,
                                product_name=line.product_name,
                                brand=line.brand,
                                quantity=line.quantity,
                                unit=line.unit,
                                shopping_detected_qty=shopping_qty,
                            )

                            if immersive_formatted and immersive_formatted.get('price_1'):
                                # Immersive results are better — use them, but backfill
                                # empty 2nd/3rd slots from shopping results
                                shopping_backup = formatted.copy()
                                _logger.info(
                                    f'[Step 2] Using immersive results for "{line.product_name}": '
                                    f'₺{immersive_formatted["price_1"]:.0f} at {immersive_formatted["store_1"]} '
                                    f'(was ₺{formatted["price_1"]:.0f} from shopping)'
                                )
                                formatted = immersive_formatted

                                # Backfill empty price slots from shopping results
                                for slot in (2, 3):
                                    if not formatted.get(f'price_{slot}') and shopping_backup.get(f'price_{slot}'):
                                        formatted[f'price_{slot}'] = shopping_backup[f'price_{slot}']
                                        formatted[f'store_{slot}'] = shopping_backup.get(f'store_{slot}', '')
                                        formatted[f'url_{slot}'] = shopping_backup.get(f'url_{slot}', '')
                                        formatted[f'detected_qty_{slot}'] = shopping_backup.get(f'detected_qty_{slot}', 0)
                                        _logger.info(
                                            f'[Step 2] Backfilled slot {slot} from shopping: '
                                            f'₺{formatted[f"price_{slot}"]:.0f} at {formatted[f"store_{slot}"]}'
                                        )
                                # Also backfill slot 1 from shopping into empty slot 2/3
                                # if immersive only had 1 seller and shopping had a different price
                                if (not formatted.get('price_2')
                                        and shopping_backup.get('price_1')
                                        and shopping_backup['store_1'] != formatted.get('store_1', '')):
                                    formatted['price_2'] = shopping_backup['price_1']
                                    formatted['store_2'] = shopping_backup.get('store_1', '')
                                    formatted['url_2'] = shopping_backup.get('url_1', '')
                                    formatted['detected_qty_2'] = shopping_backup.get('detected_qty_1', 0)
                                    _logger.info(
                                        f'[Step 2] Backfilled slot 2 from shopping best: '
                                        f'₺{formatted["price_2"]:.0f} at {formatted["store_2"]}'
                                    )
                            else:
                                _logger.info(
                                    f'[Step 2] Immersive product had no valid offers for "{line.product_name}", '
                                    f'keeping shopping results'
                                )
                    elif formatted.get('price_1') and not page_token:
                        _logger.info(
                            f'[Step 2] No page_token available for "{line.product_name}", '
                            f'keeping shopping results'
                        )

                    # Write results to the line
                    line.write({
                        'price_1': formatted.get('price_1', 0),
                        'store_1': formatted.get('store_1', ''),
                        'url_1': formatted.get('url_1', ''),
                        'price_2': formatted.get('price_2', 0),
                        'store_2': formatted.get('store_2', ''),
                        'url_2': formatted.get('url_2', ''),
                        'price_3': formatted.get('price_3', 0),
                        'store_3': formatted.get('store_3', ''),
                        'url_3': formatted.get('url_3', ''),
                        'detected_qty_1': formatted.get('detected_qty_1', 0),
                        'detected_qty_2': formatted.get('detected_qty_2', 0),
                        'detected_qty_3': formatted.get('detected_qty_3', 0),
                        'warning': formatted.get('warning', ''),
                    })

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    _logger.exception(f'Failed to search product: {line.product_name}')
                    line.write({
                        'warning': f'Search error: {str(e)[:100]}',
                    })

            # Save to price history
            self._save_price_history()

            if error_count == 0:
                self.state = 'done'
            elif success_count > 0:
                self.state = 'done'
                self.error_message = f'{error_count} product(s) had search errors. {success_count} succeeded.'
            else:
                self.state = 'error'
                self.error_message = 'All product searches failed.'

        except UserError:
            self.state = 'error'
            raise
        except Exception as e:
            self.state = 'error'
            self.error_message = str(e)
            _logger.exception('Price search failed')

    def _save_price_history(self):
        """Save current prices to history for tracking"""
        PriceHistory = self.env['totaline.price.history']
        for line in self.search_line_ids:
            if line.price_1 > 0:
                PriceHistory.create({
                    'search_id': self.id,
                    'search_line_id': line.id,
                    'product_name': line.product_name,
                    'brand': line.brand,
                    'quantity': line.quantity,
                    'unit': line.unit,
                    'detected_qty_1': line.detected_qty_1,
                    'price_1': line.price_1,
                    'store_1': line.store_1,
                    'url_1': line.url_1,
                    'price_2': line.price_2,
                    'store_2': line.store_2,
                    'price_3': line.price_3,
                    'store_3': line.store_3,
                    'search_date': self.search_date,
                })

    def _process_search_results(self, results):
        """
        Legacy: Update search lines with results from n8n.
        Kept for backward compatibility with any existing n8n workflows.
        """
        _logger.info(f'Processing search results (legacy): {type(results)}')

        # Handle nested response formats from n8n
        if isinstance(results, list) and len(results) > 0:
            first_item = results[0]
            if isinstance(first_item, dict) and 'results' in first_item:
                results = first_item.get('results', [])
                if isinstance(results, list) and len(results) > 0:
                    first_inner = results[0]
                    if isinstance(first_inner, dict) and 'results' in first_inner:
                        results = first_inner.get('results', [])

        if isinstance(results, dict):
            results = results.get('results', [results])

        if not isinstance(results, list):
            _logger.warning(f'Unexpected results format: {type(results)}')
            return

        # Build maps for matching
        line_map_by_name = {}
        line_map_by_id = {}
        for line in self.search_line_ids:
            key = line.product_name.lower().strip() if line.product_name else ''
            line_map_by_name[key] = line
            line_map_by_id[str(line.id)] = line

        for result in results:
            if not isinstance(result, dict):
                continue

            # Find matching line
            line = None
            line_id = result.get('id')
            if line_id:
                line = line_map_by_id.get(str(line_id))

            if not line:
                product_name = result.get('product', result.get('name', result.get('product_name', '')))
                product_key = product_name.lower().strip() if product_name else ''
                line = line_map_by_name.get(product_key)
                if not line:
                    continue

            # Get prices
            price_1 = result.get('price_1', 0)
            store_1 = result.get('store_1', '')
            url_1 = result.get('url_1', '')
            price_2 = result.get('price_2', 0)
            store_2 = result.get('store_2', '')
            url_2 = result.get('url_2', '')
            price_3 = result.get('price_3', 0)
            store_3 = result.get('store_3', '')
            url_3 = result.get('url_3', '')

            # Fallback to trusted_results array
            if not price_1:
                top_results = result.get('trusted_results', result.get('results', []))
                if top_results and len(top_results) > 0:
                    price_1 = top_results[0].get('totalPrice', 0)
                    store_1 = top_results[0].get('store', '')
                    url_1 = top_results[0].get('link', '')
                if top_results and len(top_results) > 1:
                    price_2 = top_results[1].get('totalPrice', 0)
                    store_2 = top_results[1].get('store', '')
                    url_2 = top_results[1].get('link', '')
                if top_results and len(top_results) > 2:
                    price_3 = top_results[2].get('totalPrice', 0)
                    store_3 = top_results[2].get('store', '')
                    url_3 = top_results[2].get('link', '')

            warning = result.get('warning', '')

            line.write({
                'price_1': price_1 or 0,
                'store_1': store_1 or '',
                'url_1': url_1 or '',
                'price_2': price_2 or 0,
                'store_2': store_2 or '',
                'url_2': url_2 or '',
                'price_3': price_3 or 0,
                'store_3': store_3 or '',
                'url_3': url_3 or '',
                'warning': warning or '',
            })

    def action_import_excel(self):
        """Open wizard to import Excel file"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Products',
            'res_model': 'totaline.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_search_id': self.id},
        }

    def action_reset_draft(self):
        """Reset to draft state"""
        self.state = 'draft'
        self.error_message = False

    def action_view_price_history(self):
        """View price history for this search"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Price History',
            'res_model': 'totaline.price.history',
            'view_mode': 'list,graph,pivot',
            'domain': [('search_id', '=', self.id)],
            'context': {'default_search_id': self.id},
        }

    def action_export_template(self):
        """Export a formatted Excel template for product import."""
        if not OPENPYXL_AVAILABLE:
            raise UserError('Excel library (openpyxl) is not installed.')

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # --- Sheet 1: Products ---
        ws = wb.active
        ws.title = 'Products'

        # Check if we have search results to include price columns
        has_results = self.state == 'done' and any(line.price_1 > 0 for line in self.search_line_ids)

        if has_results:
            headers = [
                'Product Name', 'Brand', 'Quantity', 'Package Qty', 'Items per Package',
                'Unit', 'Description',
                '1st Price', '1st Total Cost', '1st Store', '1st Store URL',
                '2nd Price', '2nd Store',
                '3rd Price', '3rd Store',
                'Warning',
            ]
        else:
            headers = ['Product Name', 'Brand', 'Quantity', 'Package Qty', 'Items per Package', 'Unit', 'Description']

        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        price_font = Font(bold=True, color='006100', size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Export existing products if available, otherwise show example
        if self.search_line_ids:
            for row_idx, line in enumerate(self.search_line_ids, 2):
                row_data = [
                    line.product_name or '',
                    line.brand or '',
                    line.total_quantity or line.quantity or 1,
                    line.package_quantity or 1,
                    line.items_per_package or 1,
                    line.unit or '',
                    line.description or '',
                ]

                if has_results:
                    row_data.extend([
                        line.price_1 or 0,
                        line.total_cost_1 or 0,
                        line.store_1 or '',
                        line.url_1 or '',
                        line.price_2 or 0,
                        line.store_2 or '',
                        line.price_3 or 0,
                        line.store_3 or '',
                        line.warning or '',
                    ])

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border

                # Highlight price columns with green
                if has_results:
                    for price_col in [8, 9]:  # 1st Price, Total Cost
                        cell = ws.cell(row=row_idx, column=price_col)
                        cell.fill = green_fill
                        cell.font = price_font
                        cell.number_format = '#,##0.00'
        else:
            example = ['Turkish Coffee 100g', 'Kurukahveci Mehmet Efendi', 12, 3, 4, 'adet', '100gr pack']
            for col_idx, val in enumerate(example, 1):
                cell = ws.cell(row=2, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = Font(italic=True, color='808080')

        # Add totals row if we have results
        if has_results and self.search_line_ids:
            total_row = len(self.search_line_ids) + 2
            ws.cell(row=total_row, column=7, value='TOTAL').font = Font(bold=True, size=12)
            ws.cell(row=total_row, column=7).border = thin_border
            total_cell = ws.cell(row=total_row, column=9,
                                 value=sum(line.total_cost_1 or 0 for line in self.search_line_ids))
            total_cell.font = Font(bold=True, color='006100', size=12)
            total_cell.fill = green_fill
            total_cell.number_format = '#,##0.00'
            total_cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 35
        if has_results:
            ws.column_dimensions['H'].width = 14  # 1st Price
            ws.column_dimensions['I'].width = 16  # Total Cost
            ws.column_dimensions['J'].width = 22  # 1st Store
            ws.column_dimensions['K'].width = 50  # URL
            ws.column_dimensions['L'].width = 14  # 2nd Price
            ws.column_dimensions['M'].width = 22  # 2nd Store
            ws.column_dimensions['N'].width = 14  # 3rd Price
            ws.column_dimensions['O'].width = 22  # 3rd Store
            ws.column_dimensions['P'].width = 40  # Warning

        # --- Sheet 2: Instructions ---
        ws2 = wb.create_sheet('Instructions')
        instructions = [
            ['Column', 'Required', 'Description', 'Example'],
            ['Product Name', 'Yes', 'Name of the product to search for', 'Turkish Coffee 100g'],
            ['Brand', 'No', 'Brand name (improves search accuracy)', 'Kurukahveci Mehmet Efendi'],
            ['Quantity', 'No', 'Total units needed (default: 1). Auto-calculated if Package Qty and Items per Package are provided.', '12'],
            ['Package Qty', 'No', 'Number of packages to order (default: 1)', '3'],
            ['Items per Package', 'No', 'Units per package (default: 1). Total = Package Qty x Items per Package.', '4'],
            ['Unit', 'No', 'Unit type: adet, kg, litre, paket, koli', 'adet'],
            ['Description', 'No', 'Additional details (size, variant, etc.)', '100gr pack'],
        ]

        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 50
        ws2.column_dimensions['D'].width = 30

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.getvalue())
        output.close()

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'product_import_template.xlsx',
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_create_purchase_order(self):
        """Create purchase order from selected prices"""
        self.ensure_one()

        if self.state != 'done':
            raise UserError('Please complete the price search first.')

        # Check if there are any results with prices
        lines_with_prices = self.search_line_ids.filtered(lambda l: l.selected_price_value > 0)
        if not lines_with_prices:
            raise UserError('No products with prices found. Please run the price search first.')

        # Get or create the vendor
        Partner = self.env['res.partner']
        vendor = Partner.search([('name', '=', 'Online Price Search')], limit=1)
        if not vendor:
            vendor = Partner.create({
                'name': 'Online Price Search',
                'supplier_rank': 1,
                'company_type': 'company',
                'comment': 'Auto-created vendor for price search purchases',
            })

        # Get product model
        Product = self.env['product.product']

        # Prepare PO lines
        po_lines = []
        notes = []
        total_selected = 0

        for line in lines_with_prices:
            # Use linked Odoo product if available, otherwise find or create
            if line.product_id:
                product = line.product_id
            else:
                product = Product.search([('name', 'ilike', line.product_name)], limit=1)

                if not product:
                    product = Product.create({
                        'name': f"{line.product_name} ({line.brand})" if line.brand else line.product_name,
                        'type': 'consu',
                        'purchase_ok': True,
                        'sale_ok': False,
                    })

            # Get selected price, store, and URL
            selected_price = line.selected_price_value  # Raw listing price from Google
            selected_store = line.selected_store_value
            selected_url = line.selected_url_value

            # Use shared pack-conversion logic for accurate PO quantities
            selected_detected = getattr(line, f'detected_qty_{line.selected_price[-1]}', 0) or 1
            packs_needed = line._get_packs_needed(selected_detected)

            # Build description
            description = f"{line.product_name}"
            if line.brand:
                description += f" - {line.brand}"
            if line.description:
                description += f" ({line.description})"
            description += f"\nStore: {selected_store}"
            if selected_url:
                description += f"\nURL: {selected_url}"
            if line.warning:
                description += f"\nNote: {line.warning}"

            # PO line: quantity = packs needed, price = listing price per pack
            po_lines.append((0, 0, {
                'product_id': product.id,
                'name': description,
                'product_qty': packs_needed,
                'price_unit': selected_price,
                'date_planned': fields.Datetime.now(),
            }))

            total_selected += selected_price * packs_needed
            notes.append(f"\u2022 {line.product_name}: \u20ba{selected_price:,.2f} from {selected_store} (Price {line.selected_price})")

        # Create the Purchase Order
        PurchaseOrder = self.env['purchase.order']
        po = PurchaseOrder.create({
            'partner_id': vendor.id,
            'date_order': fields.Datetime.now(),
            'notes': f"Created from Price Search: {self.name}\n\nProducts:\n" + "\n".join(notes),
            'order_line': po_lines,
        })

        self.message_post(
            body=f"Purchase Order <a href='#' data-oe-model='purchase.order' data-oe-id='{po.id}'>{po.name}</a> created with {len(po_lines)} products. Total: \u20ba{total_selected:,.2f}",
            message_type='notification',
        )

        return {
            'type': 'ir.actions.act_window',
            'name': 'Purchase Order',
            'res_model': 'purchase.order',
            'res_id': po.id,
            'view_mode': 'form',
            'target': 'current',
        }

    @api.model
    def _cron_run_scheduled_searches(self):
        """Cron job to run scheduled price searches"""
        now = fields.Datetime.now()
        scheduled_searches = self.search([
            ('is_scheduled', '=', True),
            ('next_scheduled_date', '<=', now),
        ])

        for search in scheduled_searches:
            try:
                _logger.info(f'Running scheduled search: {search.name}')
                search.action_search_prices()
                search.last_scheduled_run = now
                # Recompute next scheduled date
                search._compute_next_scheduled()
            except Exception as e:
                _logger.exception(f'Scheduled search failed: {search.name}')
                search.message_post(
                    body=f"Scheduled search failed: {str(e)}",
                    message_type='notification',
                )


class TotalinePriceSearchLine(models.Model):
    _name = 'totaline.price.search.line'
    _description = 'Totaline Price Search Line'
    _order = 'sequence, id'

    search_id = fields.Many2one('totaline.price.search', string='Search', ondelete='cascade')
    sequence = fields.Integer(string='Sequence', default=10)

    # Product Info (from Excel)
    product_name = fields.Char(string='Product Name', required=True)
    brand = fields.Char(string='Brand')
    quantity = fields.Float(string='Quantity', default=1)
    unit = fields.Char(string='Unit')
    description = fields.Char(string='Description')

    # Package / Unit breakdown
    package_quantity = fields.Float(string='Package Qty', default=1,
                                     help='Number of packages to order')
    items_per_package = fields.Float(string='Items/Pkg', default=1,
                                      help='Number of units per package')
    total_quantity = fields.Float(string='Total Qty', compute='_compute_total_quantity',
                                   store=True, help='package_quantity x items_per_package')

    # Detected quantities from search results (readonly)
    detected_qty_1 = fields.Float(string='Detected Qty 1', readonly=True)
    detected_qty_2 = fields.Float(string='Detected Qty 2', readonly=True)
    detected_qty_3 = fields.Float(string='Detected Qty 3', readonly=True)

    # Optimization suggestions (readonly)
    suggested_package_qty = fields.Float(string='Suggested Pkg Qty', readonly=True)
    suggested_items_per_pkg = fields.Float(string='Suggested Items/Pkg', readonly=True)
    suggested_total_price = fields.Float(string='Suggested Total Price', readonly=True)
    optimization_note = fields.Char(string='Optimization Note', readonly=True)

    # Link to existing Odoo product (optional)
    product_id = fields.Many2one('product.product', string='Odoo Product',
                                  help='Link to existing product in Odoo. Leave empty to create new.')

    # Price Selection for PO
    selected_price = fields.Selection([
        ('1', '1st Price (Best)'),
        ('2', '2nd Price'),
        ('3', '3rd Price'),
    ], string='Selected Price', default='1', help='Select which price to use for Purchase Order')

    # Price Results
    price_1 = fields.Float(string='1st Price')
    store_1 = fields.Char(string='1st Store Name')
    url_1 = fields.Char(string='1st URL')
    store_link_1 = fields.Html(string='1st Store', compute='_compute_store_links', sanitize=False)

    price_2 = fields.Float(string='2nd Price')
    store_2 = fields.Char(string='2nd Store Name')
    url_2 = fields.Char(string='2nd URL')
    store_link_2 = fields.Html(string='2nd Store', compute='_compute_store_links', sanitize=False)

    price_3 = fields.Float(string='3rd Price')
    store_3 = fields.Char(string='3rd Store Name')
    url_3 = fields.Char(string='3rd URL')
    store_link_3 = fields.Html(string='3rd Store', compute='_compute_store_links', sanitize=False)

    # Calculated: Total cost to fulfill order (listing price × packs needed)
    total_cost_1 = fields.Float(string='Total Cost', compute='_compute_total_costs', store=True)
    total_cost_2 = fields.Float(string='2nd Total Cost', compute='_compute_total_costs', store=True)
    total_cost_3 = fields.Float(string='3rd Total Cost', compute='_compute_total_costs', store=True)
    price_diff_percent = fields.Float(string='% Diff', compute='_compute_price_diff')
    warning = fields.Char(string='Warning')

    # Computed: Selected price and store for PO
    selected_price_value = fields.Float(string='Selected Price Value', compute='_compute_selected_values')
    selected_store_value = fields.Char(string='Selected Store Value', compute='_compute_selected_values')
    selected_url_value = fields.Char(string='Selected URL Value', compute='_compute_selected_values')

    # Price history link
    history_count = fields.Integer(string='History Count', compute='_compute_history_count')

    @api.depends('package_quantity', 'items_per_package', 'quantity')
    def _compute_total_quantity(self):
        for line in self:
            pkg = line.package_quantity or 1
            items = line.items_per_package or 1
            computed = pkg * items
            # If package fields are default (1×1=1) but legacy quantity > 1, use that
            if computed == 1 and (line.quantity or 0) > 1:
                line.total_quantity = line.quantity
            else:
                line.total_quantity = computed

    def _get_packs_needed(self, detected_qty):
        """Shared pack-conversion logic used by total cost computation AND PO creation.

        Returns the number of listings (packs) needed to fulfill this line's order.

        Special case: when description specifies pack contents (e.g. "8li" = 8 items
        per pack) AND detected_qty matches that number, it means the listing IS one
        package. In that case, order_qty represents number of packages, not items.
        e.g. Rulo Peçete: desc="8li", detected=8, qty=15 → 15 packs needed
        """
        self.ensure_one()
        target_qty = self.total_quantity or self.quantity or 1
        detected = detected_qty or 1

        desc_items_per_pkg = extract_items_per_pkg(self.description) if self.description else 0

        if (desc_items_per_pkg > 1
                and detected == desc_items_per_pkg
                and desc_items_per_pkg != target_qty):
            # Each listing = 1 package → need target_qty of them
            return int(target_qty)
        else:
            return math.ceil(target_qty / max(detected, 1))

    @api.depends('price_1', 'price_2', 'price_3',
                 'detected_qty_1', 'detected_qty_2', 'detected_qty_3',
                 'total_quantity', 'quantity', 'description')
    def _compute_total_costs(self):
        """Calculate total cost to fulfill the order for each price option.
        total_cost = listing_price × packs_needed
        e.g. Pril ₺69 × 4 packs = ₺276, Türk Kahvesi ₺499 × 5 packs = ₺2,495
        """
        for line in self:
            for i in range(1, 4):
                price = getattr(line, f'price_{i}', 0) or 0
                detected = getattr(line, f'detected_qty_{i}', 0) or 1

                if price > 0:
                    packs_needed = line._get_packs_needed(detected)
                    setattr(line, f'total_cost_{i}', price * packs_needed)
                else:
                    setattr(line, f'total_cost_{i}', 0)

    @api.onchange('package_quantity', 'items_per_package')
    def _onchange_package_items(self):
        """Sync legacy quantity field with total_quantity for backwards compatibility."""
        self.quantity = (self.package_quantity or 1) * (self.items_per_package or 1)

    def action_suggest_optimization(self):
        """Analyze search results and suggest optimal package/unit combination."""
        self.ensure_one()
        target_qty = self.total_quantity or self.quantity or 1
        if target_qty <= 1:
            self.optimization_note = 'No optimization needed for single unit.'
            return

        # Collect all detected quantities and prices from results
        options = []
        for i in range(1, 4):
            price = getattr(self, f'price_{i}', 0)
            detected = getattr(self, f'detected_qty_{i}', 0) or 1
            if price and price > 0:
                per_unit = price / detected if detected > 0 else price
                packs_needed = math.ceil(target_qty / detected)
                total_cost = price * packs_needed
                options.append({
                    'price_num': i,
                    'detected_qty': detected,
                    'per_unit': per_unit,
                    'packs_needed': packs_needed,
                    'total_cost': total_cost,
                })

        if not options:
            self.optimization_note = 'No price data available to optimize.'
            return

        # Find cheapest option
        best = min(options, key=lambda o: o['total_cost'])
        current_cost = self.selected_price_value or 0

        self.suggested_package_qty = best['packs_needed']
        self.suggested_items_per_pkg = best['detected_qty']
        self.suggested_total_price = best['total_cost']

        savings = current_cost - best['total_cost'] if current_cost > 0 else 0
        if savings > 0:
            self.optimization_note = (
                f"Best: {best['packs_needed']}x packs of {int(best['detected_qty'])} "
                f"(Price {best['price_num']}) = {best['total_cost']:.2f} TL "
                f"(save {savings:.2f} TL)"
            )
        else:
            self.optimization_note = (
                f"Current selection is optimal. "
                f"Best found: {best['packs_needed']}x packs of {int(best['detected_qty'])} "
                f"= {best['total_cost']:.2f} TL"
            )

    def _compute_history_count(self):
        PriceHistory = self.env['totaline.price.history']
        for line in self:
            line.history_count = PriceHistory.search_count([('search_line_id', '=', line.id)])

    def action_view_line_history(self):
        """View price history for this specific product"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': f'Price History - {self.product_name}',
            'res_model': 'totaline.price.history',
            'view_mode': 'list,graph',
            'domain': [('product_name', '=', self.product_name), ('brand', '=', self.brand or False)],
        }

    @api.depends('selected_price', 'price_1', 'price_2', 'price_3',
                 'store_1', 'store_2', 'store_3', 'url_1', 'url_2', 'url_3')
    def _compute_selected_values(self):
        for line in self:
            if line.selected_price == '2':
                line.selected_price_value = line.price_2
                line.selected_store_value = line.store_2
                line.selected_url_value = line.url_2
            elif line.selected_price == '3':
                line.selected_price_value = line.price_3
                line.selected_store_value = line.store_3
                line.selected_url_value = line.url_3
            else:  # Default to price 1
                line.selected_price_value = line.price_1
                line.selected_store_value = line.store_1
                line.selected_url_value = line.url_1

    @api.depends('store_1', 'url_1', 'store_2', 'url_2', 'store_3', 'url_3')
    def _compute_store_links(self):
        for line in self:
            for i in range(1, 4):
                url = getattr(line, f'url_{i}', '') or ''
                store = getattr(line, f'store_{i}', '') or ''
                if url and store:
                    # Escape external data (from SerpAPI) to prevent stored XSS
                    safe_store = html_escape(store)
                    safe_url = html_escape(url)
                    setattr(line, f'store_link_{i}',
                            f'<a href="{safe_url}" target="_blank">{safe_store} \u2197</a>')
                else:
                    setattr(line, f'store_link_{i}', html_escape(store) if store else '')

    @api.depends('price_1', 'price_2')
    def _compute_price_diff(self):
        for line in self:
            if line.price_1 and line.price_2:
                line.price_diff_percent = ((line.price_2 - line.price_1) / line.price_1) * 100
            else:
                line.price_diff_percent = 0


class TotalinePriceHistory(models.Model):
    _name = 'totaline.price.history'
    _description = 'Price History'
    _order = 'search_date desc'

    search_id = fields.Many2one('totaline.price.search', string='Search', ondelete='cascade')
    search_line_id = fields.Many2one('totaline.price.search.line', string='Search Line', ondelete='set null')

    # Product info snapshot
    product_name = fields.Char(string='Product Name', required=True, index=True)
    brand = fields.Char(string='Brand', index=True)
    quantity = fields.Float(string='Quantity')
    unit = fields.Char(string='Unit')

    # Price data
    price_1 = fields.Float(string='Best Price')
    store_1 = fields.Char(string='Best Store')
    url_1 = fields.Char(string='Best URL')
    price_2 = fields.Float(string='2nd Price')
    store_2 = fields.Char(string='2nd Store')
    price_3 = fields.Float(string='3rd Price')
    store_3 = fields.Char(string='3rd Store')

    # Pack detection (for accurate per-item analytics)
    detected_qty_1 = fields.Float(string='Detected Qty',
                                   help='Number of items per listing (from product title)')

    # Timing
    search_date = fields.Datetime(string='Search Date', required=True, index=True)

    # Computed for analytics
    price_per_unit = fields.Float(string='Price/Item', compute='_compute_price_per_unit', store=True)

    @api.depends('price_1', 'detected_qty_1')
    def _compute_price_per_unit(self):
        """True per-item price: listing price / items in listing.
        e.g. 8-pack of paper towels at 153 TL → 153/8 = 19.13 TL per roll.
        """
        for record in self:
            if record.price_1 and record.detected_qty_1 and record.detected_qty_1 > 0:
                record.price_per_unit = record.price_1 / record.detected_qty_1
            else:
                record.price_per_unit = record.price_1 or 0
